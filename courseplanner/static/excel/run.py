from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
import random
from datetime import datetime

# Settings
MAX_ENTRIES = 5  # Adjust this constant to set the maximum number of entries per header
TARGET_CREDIT_HRS = 120  # Target total credit hours

# Data to insert
courses = ["CS 261", "CS 240", "CS 361", "CS 227", "CS 327", "CS 482", "CS 159", "CS 412", "CS 430", 
           "CS 149", "CS 456", "CS 470", "CS 432", "CS 457", "CS 458", "CS 459", "CS 145", "CS 345", 
           "CS 455", "CS 452", "MATH 231", "MATH 245", "MATH 227", "MATH 220", "MATH 229", "MATH 318", 
           "MATH 232", "ANTH 196", "ART 224", "PHYS 326", "HIS 100", "HTH 100", "BIO 440", "BIO 160", 
           "ISAT 104", "ISCI 212", "BUS 160", "BUS 230", "BUS 400", "BUS 225", "BUS 346"]
credit_hours = [3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3]

# Create pairs of courses and credit hours, and shuffle them
course_credit_pairs = sorted(zip(courses, credit_hours), key=lambda pair: int(pair[0].split()[1]))
random.shuffle(course_credit_pairs)

def get_first_cell_of_merged_range(worksheet, row, column):
    """Find the top-left cell of the merged range containing (row, column)."""
    cell_address = f"{get_column_letter(column)}{row}"
    for merged_range in worksheet.merged_cells.ranges:
        if cell_address in merged_range:
            return merged_range.min_row, merged_range.min_col
    return row, column

def find_all_pairs_in_row(row, course_header="Course", credit_header="Credit Hrs"):
    """Find all pairs of adjacent 'Course' and 'Credit Hrs' headers in the given row."""
    pairs = []
    index = 0
    while index < len(row):
        try:
            course_col = row.index(course_header, index) + 1
            credit_col = row.index(credit_header, course_col) + 1
            pairs.append((course_col, credit_col))
            index = credit_col  # Continue searching after the current pair
        except ValueError:
            break
    return pairs

def find_all_header_pairs(worksheet, course_header="Course", credit_header="Credit Hrs", min_row=1, max_row=100):
    """Find all pairs of 'Course' and 'Credit Hrs' headers in the worksheet."""
    header_pairs = []
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row):
        pairs = find_all_pairs_in_row(row, course_header, credit_header)
        for course_col, credit_col in pairs:
            header_pairs.append((row_idx, course_col, credit_col))
    return header_pairs

def is_summer_above(worksheet, header_row, course_col):
    """Check if 'Summer' is in any cell above the specified course column."""
    for row_idx in range(1, header_row):
        if worksheet.cell(row=row_idx, column=course_col).value == "Summer":
            return True
    return False

def find_and_update_personal_info(worksheet, name, email, date):
    """Find the cells containing 'Name:', 'Email:', and 'Date:' and update the adjacent cells."""
    for row in worksheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value == "Name:":
                worksheet.cell(row=cell.row, column=cell.column + 1).value = name
            elif cell.value == "Email:":
                worksheet.cell(row=cell.row, column=cell.column + 1).value = email
            elif cell.value == "Date:":
                worksheet.cell(row=cell.row, column=cell.column + 1).value = date

def update_excel(file_path, output_path, name, email, course_header="Course", credit_header="Credit Hrs", min_row=1, max_row=100, summer=False):
    # Create a copy of the original file
    shutil.copy(file_path, output_path)

    # Load the copied workbook
    workbook = load_workbook(filename=output_path)
    worksheet = workbook.active

    # Get the current date in American format
    current_date = datetime.now().strftime("%m/%d/%Y")

    # Update the personal information and date
    find_and_update_personal_info(worksheet, name, email, current_date)

    # Find all pairs of 'Course' and 'Credit Hrs' headers
    header_pairs = find_all_header_pairs(worksheet, course_header, credit_header, min_row, max_row)
    if not header_pairs:
        raise ValueError(f"Headers '{course_header}' and/or '{credit_header}' not found in the given range.")

    # Initialize counters and totals
    total_credits = 0
    pair_index = 0

    # Fill each header pair
    for header_row, course_col, credit_col in header_pairs:
        if not summer and is_summer_above(worksheet, header_row, course_col):
            continue  # Skip this section if 'Summer' is detected above and summer is False

        row_count = 0
        while total_credits < TARGET_CREDIT_HRS and row_count < MAX_ENTRIES and pair_index < len(course_credit_pairs):
            course, credit_hr = course_credit_pairs[pair_index]

            # Find the first cell in the merged range for both headers
            course_row, course_col_final = get_first_cell_of_merged_range(worksheet, header_row + 1 + row_count, course_col)
            credit_row, credit_col_final = get_first_cell_of_merged_range(worksheet, header_row + 1 + row_count, credit_col)

            worksheet.cell(row=course_row, column=course_col_final).value = course
            worksheet.cell(row=credit_row, column=credit_col_final).value = credit_hr

            total_credits += credit_hr
            pair_index += 1
            row_count += 1

    # Save the copied workbook
    workbook.save(filename=output_path)

# Usage Example
input_file_path = 'template.xlsx'
output_file_path = 'template_copy.xlsx'
name = "John Doe"
email = "john.doe@example.com"
update_excel(input_file_path, output_file_path, name, email, min_row=11, max_row=70, summer=False)
